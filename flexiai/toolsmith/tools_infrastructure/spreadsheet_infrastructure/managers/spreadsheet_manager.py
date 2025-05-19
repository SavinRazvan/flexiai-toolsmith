# FILE: flexiai/toolsmith/tools_infrastructure/spreadsheet_infrastructure/managers/spreadsheet_manager.py

"""
spreadsheet_manager module.

Defines SpreadsheetManager, responsible for comprehensive spreadsheet operations using openpyxl, including:
- Workbook management (create, load, delete)
- Sheet management (create, rename, delete)
- Data entry (add rows, write headers, update columns)
- Data retrieval (single cell, columns, rows, filtering)
- Data analysis (summaries, structure validation, pivot tables)
- Formula operations (insert, apply to column, evaluate, remove)
- Named ranges and data transformations (define, transpose, unpivot)
- Data validation (set, remove)
- Formatting (cell styling, conditional formatting)
- Chart operations (create, update, remove)
"""

import os
import logging
import openpyxl
import pandas as pd

from typing import List, Dict, Any, Optional, Tuple, Union
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series, ScatterChart, AreaChart, BubbleChart

from flexiai.toolsmith.tools_infrastructure.spreadsheet_infrastructure.exceptions.spreadsheet_exceptions import SpreadsheetError
from flexiai.toolsmith.tools_infrastructure.spreadsheet_infrastructure.utils.file_handler import get_full_path, check_file_exists


logger = logging.getLogger(__name__)


class SpreadsheetManager:
    """
    Manages spreadsheet operations using openpyxl.
    """

    def __init__(self, file_path: str, load_workbook: bool = True):
        """
        Initialize the SpreadsheetManager.

        Args:
            file_path (str): Path to the workbook file.
            load_workbook (bool): Whether to load the workbook on init. If False, workbook is not loaded.
        """
        self.file_path = file_path
        self.workbook = None
        logger.debug(f"Initialized SpreadsheetManager with path '{self.file_path}'.")
        if load_workbook:
            self._load_workbook()

    # ------------------------------------------------------------------------------
    # 1. Factory Methods
    # ------------------------------------------------------------------------------

    @classmethod
    def create_with_defaults(cls) -> 'SpreadsheetManager':
        """
        Create a SpreadsheetManager with default path and file.

        Returns:
            SpreadsheetManager: Manager initialized with default spreadsheet.
        """
        default_path = "flexiai/toolsmith/data/spreadsheets"
        default_file_name = "example_spreadsheet.xlsx"
        full_path = get_full_path(default_path, default_file_name)
        return cls(file_path=full_path)

    # ------------------------------------------------------------------------------
    # 2. Workbook Management
    # ------------------------------------------------------------------------------

    def _load_workbook(self) -> None:
        """
        Load the workbook from disk.

        Raises:
            SpreadsheetError: If loading fails or file not found.
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            logger.info(f"Workbook '{self.file_path}' loaded successfully.")
        except FileNotFoundError:
            logger.error(f"Workbook '{self.file_path}' not found.")
            raise SpreadsheetError(f"Workbook '{self.file_path}' not found.")
        except Exception as e:
            logger.error(f"Failed to load workbook '{self.file_path}': {e}")
            raise SpreadsheetError(f"Failed to load workbook '{self.file_path}': {e}") from e

    def _ensure_workbook_loaded(self) -> None:
        """
        Ensure the workbook is loaded into memory.
        """
        if self.workbook is None:
            self._load_workbook()

    def create_workbook(self) -> None:
        """
        Create a new workbook file. Fails if file exists.

        Raises:
            SpreadsheetError: If creation fails or file already exists.
        """
        if os.path.exists(self.file_path):
            raise SpreadsheetError(f"Cannot create. File '{self.file_path}' already exists.")

        try:
            wb = Workbook()
            wb.save(self.file_path)
            self.workbook = wb  # Keep the workbook in memory
            logger.info(f"Workbook '{self.file_path}' created successfully.")
        except Exception as e:
            logger.error(f"Failed to create workbook '{self.file_path}': {e}")
            raise SpreadsheetError(f"Failed to create workbook '{self.file_path}': {e}") from e

    def delete_workbook(self) -> None:
        """
        Delete the workbook file from disk.

        Raises:
            SpreadsheetError: If deletion fails.
        """
        try:
            # If a workbook is loaded, close it
            if self.workbook:
                self.workbook.close()

            os.remove(self.file_path)
            logger.info(f"Workbook '{self.file_path}' deleted successfully.")
        except Exception as e:
            logger.error(f"Failed to delete workbook '{self.file_path}': {e}")
            raise SpreadsheetError(f"Failed to delete workbook '{self.file_path}': {e}") from e

    # ------------------------------------------------------------------------------
    # 3. Sheet Management
    # ------------------------------------------------------------------------------

    def create_sheet(self, sheet_name: str) -> None:
        """
        Create a new sheet in the workbook.

        Args:
            sheet_name (str): Title for the new sheet.

        Raises:
            SpreadsheetError: If sheet exists or creation fails.
        """
        try:
            self._ensure_workbook_loaded()
            if sheet_name in self.workbook.sheetnames:
                raise SpreadsheetError(f"Sheet '{sheet_name}' already exists.")
            self.workbook.create_sheet(title=sheet_name)
            self.workbook.save(self.file_path)
            logger.info(f"Sheet '{sheet_name}' created successfully.")
        except Exception as e:
            logger.error(f"Failed to create sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to create sheet '{sheet_name}': {e}") from e

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """
        Rename an existing sheet.

        Args:
            old_name (str): Current sheet name.
            new_name (str): Desired sheet name.

        Raises:
            SpreadsheetError: If old_name not found, new_name exists, or rename fails.
        """
        try:
            self._ensure_workbook_loaded()
            if old_name not in self.workbook.sheetnames:
                raise SpreadsheetError(f"Sheet '{old_name}' does not exist.")
            if new_name in self.workbook.sheetnames:
                raise SpreadsheetError(f"Sheet '{new_name}' already exists.")
            sheet = self.workbook[old_name]
            sheet.title = new_name
            self.workbook.save(self.file_path)
            logger.info(f"Sheet renamed from '{old_name}' to '{new_name}' successfully.")
        except Exception as e:
            logger.error(f"Failed to rename sheet '{old_name}' to '{new_name}': {e}")
            raise SpreadsheetError(f"Failed to rename sheet '{old_name}' to '{new_name}': {e}") from e

    def delete_sheet(self, sheet_name: str) -> None:
        """
        Delete a sheet from the workbook.

        Args:
            sheet_name (str): Name of the sheet to remove.

        Raises:
            SpreadsheetError: If sheet not found or deletion fails.
        """
        try:
            self._ensure_workbook_loaded()
            if sheet_name not in self.workbook.sheetnames:
                raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")
            del self.workbook[sheet_name]
            self.workbook.save(self.file_path)
            logger.info(f"Sheet '{sheet_name}' deleted successfully.")
        except Exception as e:
            logger.error(f"Failed to delete sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to delete sheet '{sheet_name}': {e}") from e

    # ------------------------------------------------------------------------------
    # 4. Data Entry Operations
    # ------------------------------------------------------------------------------

    def add_row(self, sheet_name: str, data: List[Any]) -> None:
        """
        Add a single row to a sheet.

        Args:
            sheet_name (str): Target sheet name.
            data (List[Any]): Values for the new row.

        Raises:
            SpreadsheetError: If sheet not found or append fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]
            sheet.append(data)
            self.workbook.save(self.file_path)
            logger.info(f"Row added to sheet '{sheet_name}': {data}")
        except Exception as e:
            logger.error(f"Failed to add row to sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to add row to sheet '{sheet_name}': {e}") from e

    def add_rows(self, sheet_name: str, rows: List[List[Any]]) -> None:
        """
        Add multiple rows to a sheet.

        Args:
            sheet_name (str): Target sheet name.
            rows (List[List[Any]]): List of rows to append.

        Raises:
            SpreadsheetError: If sheet not found or append fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]
            for row in rows:
                sheet.append(row)
            self.workbook.save(self.file_path)
            logger.info(f"{len(rows)} rows added to sheet '{sheet_name}'.")
        except Exception as e:
            logger.error(f"Failed to add rows to sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to add rows to sheet '{sheet_name}': {e}") from e

    def write_headers(self, sheet_name: str, headers: List[str]) -> None:
        """
        Write header row to a sheet.

        Args:
            sheet_name (str): Target sheet.
            headers (List[str]): Header labels.

        Raises:
            SpreadsheetError: If sheet not found or write fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]
            sheet.append(headers)
            self.workbook.save(self.file_path)
            logger.info(f"Headers written to sheet '{sheet_name}': {headers}")
        except Exception as e:
            logger.error(f"Failed to write headers to sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to write headers to sheet '{sheet_name}': {e}") from e

    def delete_row(self, sheet_name: str, row_id: str) -> None:
        """
        Delete a specific row by its 1-based index.

        Args:
            sheet_name (str): Target sheet name.
            row_id (str): Row number as string.

        Raises:
            SpreadsheetError: If invalid row_id or deletion fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]
            row_number = int(row_id)
            if row_number < 1 or row_number > sheet.max_row:
                raise SpreadsheetError(f"Row number '{row_number}' is out of range in sheet '{sheet_name}'.")
            sheet.delete_rows(row_number)
            self.workbook.save(self.file_path)
            logger.info(f"Row '{row_number}' deleted from sheet '{sheet_name}'.")
        except ValueError:
            logger.error(f"Invalid row_id '{row_id}'. Must be an integer.")
            raise SpreadsheetError(f"Invalid row_id '{row_id}'. Must be an integer.")
        except Exception as e:
            logger.error(f"Failed to delete row '{row_id}' from sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to delete row '{row_id}' from sheet '{sheet_name}': {e}") from e

    def update_column(
        self,
        sheet_name: str,
        column_identifier: Union[str, int],
        new_data: List[Any],
        skip_header: bool = True,
        has_headers: bool = True
    ) -> int:
        """
        Update values in a column.

        Args:
            sheet_name (str): Target sheet name.
            column_identifier (Union[str,int]): Letter, index, or header name.
            new_data (List[Any]): New cell values.
            skip_header (bool): If True, skip first row.
            has_headers (bool): If True, treat first row as headers.

        Returns:
            int: Number of cells updated.

        Raises:
            SpreadsheetError: If resolution fails or write fails.
        """
        self._ensure_workbook_loaded()
        sheet = self.workbook[sheet_name]
        
        # Use the new resolver instead of _column_index
        col_idx = self._resolve_column_identifier(sheet_name, column_identifier, has_headers)

        # If skip_header, start from row=2, else row=1
        start_row = 2 if skip_header else 1

        rows_updated = 0
        for value in new_data:
            cell = sheet.cell(row=start_row + rows_updated, column=col_idx + 1)  
            cell.value = value
            rows_updated += 1

        self.workbook.save(self.file_path)
        logger.info(
            f"Column '{column_identifier}' (resolved to index {col_idx}) updated successfully. "
            f"Rows updated: '{rows_updated}', on sheet: '{sheet_name}' with data: '{new_data}'."
        )
        return rows_updated


    # ------------------------------------------------------------------------------
    # 5. Data Retrieval Operations
    # ------------------------------------------------------------------------------

    def retrieve_cell(
        self,
        sheet_name: str,
        column_identifier: Union[str, int],
        row_number: int,
        skip_header: bool = False,
        has_headers: bool = True
    ) -> Any:
        """
        Retrieve a single cell value.

        Args:
            sheet_name (str): Target sheet name.
            column_identifier (Union[str,int]): Letter, index, or header name.
            row_number (int): 1-based row number.
            skip_header (bool): If True and row_number==1, maps to row 2.
            has_headers (bool): If True, interpret headers row.

        Returns:
            Any: Cell value.

        Raises:
            SpreadsheetError: If out of range or resolution fails.
        """
        self._ensure_workbook_loaded()
        if sheet_name not in self.workbook.sheetnames:
            raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")
        
        sheet = self.workbook[sheet_name]

        # If skip_header, row_number=1 => actual row=2
        actual_row = row_number + 1 if (skip_header and row_number == 1) else row_number
        if actual_row < 1 or actual_row > sheet.max_row:
            raise SpreadsheetError(f"Row '{actual_row}' out of range in sheet '{sheet_name}'.")

        col_idx = self._resolve_column_identifier(sheet_name, column_identifier, has_headers)
        # col_idx is 0-based => add +1 for openpyxl
        cell_obj = sheet.cell(row=actual_row, column=col_idx + 1)
        value = cell_obj.value
        logger.info(
            f"retrieve_cell -> Column '{column_identifier}', row {row_number} (actual={actual_row}), "
            f"skip_header={skip_header}, sheet='{sheet_name}' => {value}"
        )
        return value


    def retrieve_column(
        self,
        sheet_name: str,
        column_identifier: Union[str, int],
        skip_header: bool = False,
        has_headers: bool = True
    ) -> List[Any]:
        """
        Retrieve all values from a column.

        Args:
            sheet_name (str): Target sheet name.
            column_identifier (Union[str,int]): Letter, index, or header name.
            skip_header (bool): If True, skip first row.
            has_headers (bool): If True, treat first row as headers.

        Returns:
            List[Any]: Column values.

        Raises:
            SpreadsheetError: If sheet not found or resolution fails.
        """
        self._ensure_workbook_loaded()
        if sheet_name not in self.workbook.sheetnames:
            raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")

        sheet = self.workbook[sheet_name]

        # Resolve column to 0-based index
        col_idx = self._resolve_column_identifier(sheet_name, column_identifier, has_headers)

        # If skip_header => start from row=2, else from row=1
        start_row = 2 if skip_header else 1
        end_row = sheet.max_row

        result = []
        for row_num in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row_num, column=col_idx + 1).value
            result.append(cell_value)

        logger.info(
            f"retrieve_column -> Column '{column_identifier}' (index {col_idx}) from '{sheet_name}', "
            f"skip_header={skip_header}, total rows returned={len(result)}."
        )
        return result


    def filter_rows(
        self,
        sheet_name: str,
        column_identifier: Union[str, int],
        condition_type: str,
        condition_value: str,
        skip_header: bool = True,
        has_headers: bool = True
    ) -> List[List[Any]]:
        """
        Filter rows by a condition on a column.

        Args:
            sheet_name (str): Target sheet name.
            column_identifier (Union[str,int]): Letter, index, or header name.
            condition_type (str): 'equals','greater_than', etc.
            condition_value (str): Value to compare.
            skip_header (bool): If True, skip first row.
            has_headers (bool): If True, treat headers row.

        Returns:
            List[List[Any]]: Matching rows.

        Raises:
            SpreadsheetError: If sheet not found or invalid condition.
        """
        self._ensure_workbook_loaded()
        if sheet_name not in self.workbook.sheetnames:
            raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")

        sheet = self.workbook[sheet_name]
        col_idx = self._resolve_column_identifier(sheet_name, column_identifier, has_headers)

        condition_func = self._build_condition_func(condition_type, condition_value)
        filtered_rows = []

        # If skip_header => start from row=2, else from row=1
        start_row = 2 if skip_header else 1

        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            # row is a tuple of cell values
            cell_value = row[col_idx] if col_idx < len(row) else None
            if condition_func(cell_value):
                filtered_rows.append(list(row))

        logger.info(
            f"filter_rows -> Filtered by '{condition_type}'='{condition_value}' "
            f"on column '{column_identifier}' in sheet '{sheet_name}'. "
            f"Found {len(filtered_rows)} matching rows."
        )
        return filtered_rows


    def retrieve_rows(
        self,
        sheet_name: str,
        start_row: int = 1,
        max_rows: int = 20,
        skip_header: bool = False
    ) -> List[List[Any]]:
        """
        Retrieve a block of rows.

        Args:
            sheet_name (str): Target sheet name.
            start_row (int): 1-based starting row.
            max_rows (int): Maximum rows to fetch.
            skip_header (bool): If True and start_row<=1, skip header row.

        Returns:
            List[List[Any]]: Retrieved rows.

        Raises:
            SpreadsheetError: If sheet not found.
        """
        self._ensure_workbook_loaded()
        if sheet_name not in self.workbook.sheetnames:
            raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")

        sheet = self.workbook[sheet_name]

        # If skip_header and start_row=1, effectively start at row=2
        actual_start = start_row + 1 if (skip_header and start_row <= 1) else start_row
        end_row = actual_start + max_rows - 1

        rows_data = []
        for idx, row in enumerate(
            sheet.iter_rows(min_row=actual_start, max_row=end_row, values_only=True),
            start=actual_start
        ):
            rows_data.append(list(row))

        logger.info(
            f"retrieve_rows -> From sheet '{sheet_name}', "
            f"requested rows {actual_start} to {end_row} (skip_header={skip_header}). "
            f"Returned {len(rows_data)} rows."
        )
        return rows_data


    def retrieve_row(
        self,
        sheet_name: str,
        row_id: int,
        skip_header: bool = False
    ) -> List[Any]:
        """
        Retrieve a single row.

        Args:
            sheet_name (str): Target sheet name.
            row_id (int): 1-based row number.
            skip_header (bool): If True and row_id==1, maps to row 2.

        Returns:
            List[Any]: Row values.

        Raises:
            SpreadsheetError: If sheet not found or out of range.
        """
        self._ensure_workbook_loaded()
        if sheet_name not in self.workbook.sheetnames:
            raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")

        sheet = self.workbook[sheet_name]

        # If row 1 is a header and skip_header=True, then retrieving row_id=1 means row #2
        actual_row = row_id + 1 if (skip_header and row_id == 1) else row_id

        if actual_row < 1 or actual_row > sheet.max_row:
            raise SpreadsheetError(
                f"Row number '{actual_row}' is out of range in sheet '{sheet_name}'."
            )

        # gather that row
        row_data = [cell.value for cell in sheet[actual_row]]

        logger.info(
            f"retrieve_row -> Row {row_id} (actual={actual_row}), skip_header={skip_header}, "
            f"sheet='{sheet_name}' -> {row_data}"
        )
        return row_data


    # ------------------------------------------------------------------------------
    # 6. Data Analysis Operations
    # ------------------------------------------------------------------------------

    def generate_spreadsheet_summary(self) -> Dict[str, Any]:
        """
        Generate row/column counts for each sheet.

        Returns:
            Dict[str, Any]: {sheet_name: {'rows': int, 'columns': int}, ...}

        Raises:
            SpreadsheetError: If generation fails.
        """
        try:
            self._ensure_workbook_loaded()
            summary = {}
            for sheet_name in self.workbook.sheetnames:
                sheet_obj = self.workbook[sheet_name]
                summary[sheet_name] = {
                    "rows": sheet_obj.max_row,
                    "columns": sheet_obj.max_column
                }
            logger.info(f"Spreadsheet summary generated successfully for '{self.file_path}'.")
            return summary
        except Exception as e:
            logger.error(f"Failed to generate spreadsheet summary: {e}")
            raise SpreadsheetError(f"Failed to generate spreadsheet summary: {e}") from e


    def validate_spreadsheet_structure(
        self,
        required_sheets: List[str],
        required_headers: Dict[str, List[str]]
    ) -> bool:
        """
        Validate that required sheets and headers exist.

        Args:
            required_sheets (List[str]): Sheet names that must be present.
            required_headers (Dict[str,List[str]]): {sheet_name: [header1,...], ...}

        Returns:
            bool: True if valid.

        Raises:
            SpreadsheetError: If any sheet or header is missing.
        """
        try:
            self._ensure_workbook_loaded()

            # Check for required sheets
            for sheet_name in required_sheets:
                if sheet_name not in self.workbook.sheetnames:
                    error_msg = f"Sheet '{sheet_name}' does not exist."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)

            # Check for required headers in each sheet
            for sheet_name, headers in required_headers.items():
                if sheet_name not in self.workbook.sheetnames:
                    error_msg = f"Sheet '{sheet_name}' does not exist for header validation."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)
                sheet_obj = self.workbook[sheet_name]
                actual_headers = list(next(sheet_obj.iter_rows(min_row=1, max_row=1, values_only=True), []))
                if actual_headers != headers:
                    error_msg = f"Headers mismatch in sheet '{sheet_name}'. Expected {headers}, got {actual_headers}."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)

            logger.info(f"Spreadsheet structure validation passed for '{self.file_path}'.")
            return True
        except SpreadsheetError:
            raise
        except Exception as e:
            logger.error(f"Failed to validate spreadsheet structure: {e}")
            raise SpreadsheetError(f"Failed to validate spreadsheet structure: {e}") from e


    def retrieve_multiple_sheets_summary(
        self,
        files_list: List[Dict[str, str]],
        default_path: str,
        default_file_name: str
    ) -> Dict[str, Any]:
        """
        Generate summaries for multiple spreadsheet files.

        Args:
            files_list (List[Dict[str,str]]): [{'path':..., 'file_name':...}, ...]
            default_path (str): Fallback directory.
            default_file_name (str): Fallback file name.

        Returns:
            Dict[str,Any]: {file_name: {'status': bool, 'message': str, 'result': summary_or_None}, ...}
        """
        summaries = {}
        for file in files_list:
            path = file.get('path', default_path)
            file_name = file.get('file_name', default_file_name)
            full_path = get_full_path(path, file_name)
            try:
                check_file_exists(path, file_name)
                manager = SpreadsheetManager(file_path=full_path)
                summary = manager.generate_spreadsheet_summary()
                message = f"Summary generated successfully for '{file_name}'."
                summaries[file_name] = {
                    "status": True,
                    "message": message,
                    "result": summary
                }
            except SpreadsheetError as e:
                summaries[file_name] = {
                    "status": False,
                    "message": str(e),
                    "result": None
                }
            except Exception as e:
                summaries[file_name] = {
                    "status": False,
                    "message": f"Unexpected error: {str(e)}",
                    "result": None
                }
        logger.info("Retrieved multiple sheets summaries.")
        return summaries


    def create_pivot_table(
        self,
        sheet_name: str,
        source_data: str,
        destination: str,
        rows: Optional[List[str]] = None,
        columns: Optional[List[str]] = None,
        values: Optional[List[Dict[str, str]]] = None,
        filters: Optional[List[str]] = None,
        page_fields: Optional[List[str]] = None,
        report_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Create and insert a pandas pivot table into the workbook.

        Args:
            sheet_name (str): Sheet to source/paste table.
            source_data (str): e.g. 'Data!A1:D100'
            destination (str): e.g. 'Summary!A1'
            rows, columns, values, filters, page_fields: pivot config
            report_name (str): Optional title above table.

        Returns:
            Dict[str,Any]: {status, pivot_table_location, report_name, sheet_name}

        Raises:
            SpreadsheetError: on invalid input or failure at any step.
        """
        # Step 0: Validate required parameters
        if not all([source_data, destination]):
            raise SpreadsheetError("Both 'source_data' and 'destination' parameters are required for creating a pivot table.")

        # Step 1: Parse and validate source_data
        try:
            source_sheet_name, source_range = source_data.split('!')
            source_sheet = self.workbook[source_sheet_name]
            
            # Validate source range
            min_col, min_row, max_col, max_row = range_boundaries(source_range)
            logger.debug(f"Extracting data from '{source_range}' (Columns {min_col}-{max_col}, Rows {min_row}-{max_row})")
            
            # Extract data using openpyxl
            source_data_rows = []
            for row in source_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
                source_data_rows.append(list(row))
            
            if not source_data_rows:
                error_msg = f"No data found in the source data range '{source_data}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            headers = source_data_rows[0]
            if not headers:
                error_msg = f"No headers found in the source data range '{source_data}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            df = pd.DataFrame(source_data_rows[1:], columns=headers)
            if df.empty:
                error_msg = f"No data found in the source data range '{source_data}' after excluding headers."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            logger.debug(f"Data extracted successfully with {len(df)} records.")
        except ValueError:
            error_msg = f"Invalid source_data format: '{source_data}'. Expected format 'SheetName!A1:D100'."
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)
        except KeyError:
            error_msg = f"Source sheet '{source_sheet_name}' does not exist."
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)
        except Exception as e:
            error_msg = f"Unexpected error parsing source_data: {e}"
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)

        # Step 2: Validate that specified fields exist
        try:
            all_fields = []
            if rows:
                all_fields.extend(rows)
            if columns:
                all_fields.extend(columns)
            if values:
                all_fields.extend([v['field'] for v in values])
            if filters:
                all_fields.extend(filters)
            if page_fields:
                all_fields.extend(page_fields)

            missing_fields = [field for field in all_fields if field not in df.columns]
            if missing_fields:
                error_msg = f"The following fields are missing in source data: {missing_fields}"
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            logger.debug("All specified fields exist in the source data.")
        except Exception as e:
            error_msg = f"Failed to validate pivot table fields: {e}"
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)

        # Step 3: Apply filters if any
        if filters:
            try:
                for filter_col in filters:
                    if filter_col not in df.columns:
                        error_msg = f"Filter column '{filter_col}' does not exist in the source data."
                        logger.error(error_msg)
                        raise SpreadsheetError(error_msg)
                    df = df[df[filter_col].notna()]
                if df.empty:
                    error_msg = f"No data remaining after applying filters on columns: {filters}."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)
                logger.debug(f"Data after applying filters: {len(df)} records remaining.")
            except Exception as e:
                error_msg = f"Failed to apply filters: {e}"
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

        # Step 4: Apply page_fields as additional filters if any
        if page_fields:
            try:
                for page_field in page_fields:
                    if page_field not in df.columns:
                        error_msg = f"Page field '{page_field}' does not exist in the source data."
                        logger.error(error_msg)
                        raise SpreadsheetError(error_msg)
                    unique_values = df[page_field].unique()
                    if len(unique_values) > 0:
                        # For demonstration, select the first unique value
                        selected_value = unique_values[0]
                        df = df[df[page_field] == selected_value]
                        logger.debug(f"Applied page_field filter: {page_field}='{selected_value}'")
                    else:
                        error_msg = f"No unique values found in page field '{page_field}'."
                        logger.error(error_msg)
                        raise SpreadsheetError(error_msg)
                if df.empty:
                    error_msg = f"No data remaining after applying page fields: {page_fields}."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)
                logger.debug(f"Data after applying page fields: {len(df)} records remaining.")
            except Exception as e:
                error_msg = f"Failed to apply page_fields: {e}"
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

        # Step 5: Create the pivot table using pandas
        try:
            aggfunc_dict = {}
            values_fields = []
            if values:
                for v in values:
                    field = v['field']
                    aggfunc = v['aggfunc']
                    if aggfunc not in ['sum', 'mean', 'count', 'min', 'max', 'median', 'std']:
                        error_msg = f"Unsupported aggregation function '{aggfunc}' for field '{field}'."
                        logger.error(error_msg)
                        raise SpreadsheetError(error_msg)
                    aggfunc_dict[field] = aggfunc
                    values_fields.append(field)

            pivot_table = pd.pivot_table(
                df,
                index=rows if rows else None,
                columns=columns if columns else None,
                values=values_fields if values_fields else None,
                aggfunc=aggfunc_dict if aggfunc_dict else None,
                fill_value=0,
                dropna=False
            )

            if pivot_table.empty:
                error_msg = "Pivot table is empty. Check the pivot table configuration and source data."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            logger.debug(f"Pivot table created with shape {pivot_table.shape}.")
        except Exception as e:
            error_msg = f"Failed to create pivot table with provided configuration: {e}"
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)

        # Step 6: Write the pivot table back to the destination range
        try:
            dest_sheet_name, dest_cell = destination.split('!')
            if dest_sheet_name not in self.workbook.sheetnames:
                error_msg = f"Destination sheet '{dest_sheet_name}' does not exist."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            dest_sheet = self.workbook[dest_sheet_name]

            # Set the report name first to avoid overwriting
            if report_name:
                dest_sheet[dest_cell] = report_name
                dest_sheet[dest_cell].font = Font(bold=True, size=14)
                # Extract row number from destination cell
                start_row_str = ''.join(filter(str.isdigit, dest_cell))
                if not start_row_str:
                    error_msg = f"Invalid destination cell '{dest_cell}'. No row number found."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)
                start_row = int(start_row_str) + 1  # Start writing data below the report name
                logger.debug(f"Report name set at '{dest_cell}'. Starting pivot table at row {start_row}.")
            else:
                # Extract row number from destination cell
                start_row_str = ''.join(filter(str.isdigit, dest_cell))
                if not start_row_str:
                    error_msg = f"Invalid destination cell '{dest_cell}'. No row number found."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)
                start_row = int(start_row_str)
                logger.debug(f"Starting pivot table at row {start_row}.")

            # Determine starting column
            dest_col_letter = ''.join(filter(str.isalpha, dest_cell))
            if not dest_col_letter:
                error_msg = f"Invalid destination cell '{dest_cell}'. No column letter found."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            dest_col_index = column_index_from_string(dest_col_letter)
            logger.debug(f"Starting pivot table at column '{dest_col_letter}' (index {dest_col_index}).")

            # Convert pivot table DataFrame to list of lists
            pivot_data = pivot_table.reset_index().fillna('').values.tolist()

            # Write data to the destination sheet starting from start_row
            for r_idx, row in enumerate(pivot_data, start=start_row):
                for c_idx, value in enumerate(row, start=dest_col_index):
                    dest_sheet.cell(row=r_idx, column=c_idx, value=value)
            logger.debug(f"Pivot table data written successfully to '{destination}'.")

            # Save the workbook
            self.workbook.save(self.file_path)
            logger.info(f"Pivot table '{report_name}' created successfully at '{destination}' in sheet '{dest_sheet_name}'.")

            return {
                "status": True,
                "pivot_table_location": destination,
                "report_name": report_name,
                "sheet_name": dest_sheet_name
            }
        except ValueError:
            error_msg = f"Invalid destination format: '{destination}'. Expected format 'SheetName!A1'."
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)
        except Exception as e:
            error_msg = f"Failed to write pivot table to destination '{destination}': {e}"
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)

    # ------------------------------------------------------------------------------
    # 7. Formula Operations
    # ------------------------------------------------------------------------------

    def insert_formula(self, sheet_name: str, cell: str, formula: str) -> None:
        """
        Insert a formula into a cell.

        Args:
            sheet_name (str): Target sheet.
            cell (str): Cell reference (e.g. 'B2').
            formula (str): Formula string starting with '='.

        Raises:
            SpreadsheetError: If sheet not found, invalid cell, or formula bad.
        """
        try:
            self._ensure_workbook_loaded()
            
            # Validate sheet existence
            if sheet_name not in self.workbook.sheetnames:
                error_msg = f"Sheet '{sheet_name}' does not exist in the workbook."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            try:
                column, row = coordinate_from_string(cell)
            except ValueError:
                error_msg = f"Invalid cell reference: '{cell}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            # Validate formula
            if not formula.startswith('='):
                error_msg = f"Invalid formula '{formula}'. Formulas must start with '='."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            sheet = self.workbook[sheet_name]
            sheet[cell] = formula
            self.workbook.save(self.file_path)
            logger.info(f"Inserted formula '{formula}' into cell '{cell}' in sheet '{sheet_name}'.")
        except SpreadsheetError:
            # Re-raise SpreadsheetError without modification
            raise
        except Exception as e:
            logger.error(f"Failed to insert formula into cell '{cell}' in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to insert formula into cell '{cell}' in sheet '{sheet_name}': {e}") from e



    def apply_formula_to_column(self, sheet_name: str, column: str, formula_template: str, start_row: Optional[int] = 1) -> int:
        """
        Apply a formula template down a column.

        Args:
            sheet_name (str): Target sheet.
            column (str): Column letter.
            formula_template (str): Contains '{row}' placeholder.
            start_row (int): Starting row (1-based).

        Returns:
            int: Rows updated.

        Raises:
            SpreadsheetError: On invalid parameters or failure.
        """
        try:
            self._ensure_workbook_loaded()
            
            # Validate sheet existence
            if sheet_name not in self.workbook.sheetnames:
                error_msg = f"Sheet '{sheet_name}' does not exist in the workbook."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            # Validate column name
            from openpyxl.utils import column_index_from_string
            try:
                column_index = column_index_from_string(column)
            except ValueError:
                error_msg = f"Invalid column name: '{column}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            # Validate formula_template contains '{row}'
            if '{row}' not in formula_template:
                error_msg = "Formula template must contain '{row}' placeholder."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            # Validate start_row
            if not isinstance(start_row, int) or start_row < 1:
                error_msg = f"Start row must be a positive integer. Provided: {start_row}."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            sheet = self.workbook[sheet_name]
            
            # Check if start_row exceeds max_row
            if start_row > sheet.max_row:
                error_msg = f"Start row {start_row} exceeds the number of rows in sheet '{sheet_name}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            rows_updated = 0
            for row in range(start_row, sheet.max_row + 1):
                cell_ref = f"{column}{row}"
                try:
                    formula = formula_template.format(row=row)
                except KeyError as ke:
                    error_msg = f"Missing placeholder in formula template: {ke}"
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg) from ke
                sheet[cell_ref] = formula
                rows_updated += 1
            self.workbook.save(self.file_path)
            logger.info(f"Applied formula template '{formula_template}' to column '{column}' starting at row {start_row}. Rows updated: {rows_updated}.")
            return rows_updated
        except SpreadsheetError:
            # Re-raise SpreadsheetError without modification
            raise
        except Exception as e:
            logger.error(f"Failed to apply formula to column '{column}' in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to apply formula to column '{column}' in sheet '{sheet_name}': {e}") from e


    def evaluate_formula(self, sheet_name: str, cell: str) -> Any:
        """
        Retrieve a cell’s value if not a formula.

        Args:
            sheet_name (str): Target sheet.
            cell (str): Cell reference.

        Returns:
            Any: Cell value.

        Raises:
            SpreadsheetError: If cell contains a formula or failure.
        """
        try:
            self._ensure_workbook_loaded()
            
            # Validate sheet existence
            if sheet_name not in self.workbook.sheetnames:
                error_msg = f"Sheet '{sheet_name}' does not exist in the workbook."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            # Validate cell reference
            from openpyxl.utils import coordinate_from_string
            try:
                column, row = coordinate_from_string(cell)
            except ValueError:
                error_msg = f"Invalid cell reference: '{cell}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            sheet = self.workbook[sheet_name]
            value = sheet[cell].value
            if isinstance(value, str) and value.startswith('='):
                error_msg = "Cannot evaluate formula using openpyxl. Must use Excel or another tool."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            logger.info(f"Retrieved value from cell '{cell}' in sheet '{sheet_name}': {value}")
            return value
        except SpreadsheetError:
            # Re-raise SpreadsheetError without modification
            raise
        except Exception as e:
            logger.error(f"Failed to retrieve value from cell '{cell}' in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to retrieve value from cell '{cell}' in sheet '{sheet_name}': {e}") from e


    def remove_formula(self, sheet_name: str, cell: str) -> None:
        """
        Remove a formula from a cell.

        Args:
            sheet_name (str): Target sheet.
            cell (str): Cell reference (e.g. 'C3').

        Raises:
            SpreadsheetError: If no formula or failure.
        """
        try:
            self._ensure_workbook_loaded()
            
            # Validate sheet existence
            if sheet_name not in self.workbook.sheetnames:
                error_msg = f"Sheet '{sheet_name}' does not exist in the workbook."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            # Validate cell reference
            from openpyxl.utils import coordinate_from_string
            try:
                column, row = coordinate_from_string(cell)
            except ValueError:
                error_msg = f"Invalid cell reference: '{cell}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            sheet = self.workbook[sheet_name]
            value = sheet[cell].value
            if not (isinstance(value, str) and value.startswith('=')):
                error_msg = f"Cell '{cell}' in sheet '{sheet_name}' does not contain a formula."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            sheet[cell].value = None
            self.workbook.save(self.file_path)
            logger.info(f"Removed formula from cell '{cell}' in sheet '{sheet_name}'.")
        except SpreadsheetError:
            # Re-raise SpreadsheetError without modification
            raise
        except Exception as e:
            logger.error(f"Failed to remove formula from cell '{cell}' in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to remove formula from cell '{cell}' in sheet '{sheet_name}': {e}") from e


    def define_named_range(self, sheet_name: str, range_name: str, cell_range: str) -> None:
        """
        Define or overwrite a named range.

        Args:
            sheet_name (str): Target sheet.
            range_name (str): Name for the range.
            cell_range (str): Range string (e.g. 'A1:C3').

        Raises:
            SpreadsheetError: On invalid sheet or range.
        """
        try:
            self._ensure_workbook_loaded()
            
            # Validate sheet existence
            if sheet_name not in self.workbook.sheetnames:
                error_msg = f"Sheet '{sheet_name}' does not exist in the workbook."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            sheet = self.workbook[sheet_name]

            # Validate and parse cell_range
            from openpyxl.utils import coordinate_from_string
            try:
                if ':' in cell_range:
                    start_cell, end_cell = cell_range.split(':')
                    coordinate_from_string(start_cell)  # Validates start_cell
                    coordinate_from_string(end_cell)    # Validates end_cell
                else:
                    coordinate_from_string(cell_range)  # Validates single cell
            except ValueError:
                error_msg = f"Invalid cell range: '{cell_range}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)
            
            # Remove all existing named ranges with the same name
            existing_named_ranges = [dn for dn in self.workbook.defined_names.definedName if dn.name == range_name]
            for dn in existing_named_ranges:
                self.workbook.defined_names.remove(dn)
                logger.debug(f"Existing named range '{range_name}' removed.")
            
            # Define the new named range
            dn = DefinedName(name=range_name, attr_text=f"'{sheet_name}'!{cell_range}")
            self.workbook.defined_names.append(dn)
            self.workbook.save(self.file_path)
            logger.info(f"Named range '{range_name}' defined as '{cell_range}' in sheet '{sheet_name}'.")
        except SpreadsheetError:
            # Re-raise SpreadsheetError without modification
            raise
        except Exception as e:
            logger.error(f"Failed to define named range '{range_name}' in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to define named range '{range_name}' in sheet '{sheet_name}': {e}") from e


    # ------------------------------------------------------------------------------
    # 8. Data Transformation Operations
    # ------------------------------------------------------------------------------

    def transpose_data(self, source_range: str, destination_range: str) -> None:
        """
        Transpose one range into another.

        Args:
            source_range (str): e.g. 'Sheet1!A1:C3'
            destination_range (str): e.g. 'Sheet2!A1'

        Raises:
            SpreadsheetError: On invalid ranges or failure.
        """
        try:
            self._ensure_workbook_loaded()
            
            # Parse and validate source range
            source_sheet, source_cells = self._parse_range(source_range)
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(source_cells)
            
            # Parse and validate destination range
            dest_sheet, dest_cells = self._parse_range(destination_range)
            dest_col_letter, dest_row = coordinate_from_string(dest_cells)
            dest_col_index = column_index_from_string(dest_col_letter)
            
            # Extract source data
            source_data = []
            for row in source_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
                source_data.append(list(row))
            
            # Transpose data
            transposed_data = list(map(list, zip(*source_data)))
            
            # Write transposed data to destination
            for idx, row in enumerate(transposed_data, start=dest_row):
                for jdx, value in enumerate(row, start=dest_col_index):
                    dest_sheet.cell(row=idx, column=jdx).value = value
            
            self.workbook.save(self.file_path)
            logger.info(f"Data transposed from '{source_range}' to '{destination_range}'.")
        except Exception as e:
            logger.error(f"Failed to transpose data from '{source_range}' to '{destination_range}': {e}")
            raise SpreadsheetError(f"Failed to transpose data from '{source_range}' to '{destination_range}': {e}") from e     


    def unpivot_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Convert wide-format to long-format.

        Args:
            sheet_name (str): Source sheet.

        Returns:
            List[Dict[str, Any]]: Unpivoted records.

        Raises:
            SpreadsheetError: On failure.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]
            headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            unpivoted_data = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                identifier = row[0]
                for idx, cell_value in enumerate(row[1:], start=1):
                    unpivoted_data.append({
                        "Identifier": identifier,
                        "Attribute": headers[idx],
                        "Value": cell_value
                    })

            self.workbook.save(self.file_path)
            logger.info(f"Data in sheet '{sheet_name}' unpivoted successfully.")
            return unpivoted_data
        except Exception as e:
            logger.error(f"Failed to unpivot data in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to unpivot data in sheet '{sheet_name}': {e}") from e

    # ------------------------------------------------------------------------------
    # 9. Data Validation Operations (set_data_validation, remove_data_validation)
    # ------------------------------------------------------------------------------

    def set_data_validation(self, sheet_name: str, validation_rules: Dict[str, Any]) -> None:
        """
        Sets data validation rules for a specific range in a sheet.

        Args:
            sheet_name (str): Name of the target sheet.
            validation_rules (Dict[str, Any]): Dictionary containing validation parameters.
                Example:
                {
                    "range": "A1:A10",
                    "type": "list",
                    "operator": "equal",
                    "formula1": '"Option1,Option2,Option3"',
                    "allow_blank": True,
                    "showDropDown": True,
                    "errorTitle": "Invalid Entry",
                    "error": "Please select a value from the list."
                }

        Raises:
            SpreadsheetError: If validation rules are invalid or application fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]

            # Extract validation parameters
            dv_type = validation_rules.get("type")
            operator = validation_rules.get("operator")
            formula1 = validation_rules.get("formula1")
            formula2 = validation_rules.get("formula2", None)
            allow_blank = validation_rules.get("allow_blank", True)
            show_error_message = validation_rules.get("show_error_message", True)
            error_title = validation_rules.get("errorTitle", "Invalid Entry")
            error = validation_rules.get("error", "The value entered is not valid.")
            show_drop_down = validation_rules.get("showDropDown", True)

            if not dv_type or not operator or not formula1:
                raise SpreadsheetError("Validation rules must include 'type', 'operator', and 'formula1'.")

            # Validate and parse the target range
            target_range = validation_rules.get("range")
            if not target_range:
                raise SpreadsheetError("Validation rules must include a 'range'.")

            if ':' in target_range:
                start_cell, end_cell = target_range.split(':')
                coordinate_from_string(start_cell)  # Validates start_cell
                coordinate_from_string(end_cell)    # Validates end_cell
            else:
                coordinate_from_string(target_range)  # Validates single cell

            # Create DataValidation object
            dv = DataValidation(
                type=dv_type,
                operator=operator,
                formula1=formula1,
                formula2=formula2,
                allow_blank=allow_blank,
                showErrorMessage=show_error_message,
                errorTitle=error_title,
                error=error,
                showDropDown=show_drop_down
            )

            # Apply data validation to the target range
            dv.add(target_range)
            sheet.add_data_validation(dv)

            self.workbook.save(self.file_path)
            logger.info(f"Data validation set for range '{target_range}' in sheet '{sheet_name}': {validation_rules}")
        except Exception as e:
            logger.error(f"Failed to set data validation for range '{validation_rules.get('range')}' in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to set data validation for range '{validation_rules.get('range')}' in sheet '{sheet_name}': {e}") from e


    def remove_data_validation(self, sheet_name: str, range_to_remove: Optional[str] = None) -> None:
        """
        Removes data validation rules from a specific range or the entire sheet.

        Args:
            sheet_name (str): Name of the target sheet.
            range_to_remove (str, optional): Specific range to remove validations from (e.g., "A1:A10").
                If None, removes all data validations from the sheet.

        Raises:
            SpreadsheetError: If removal fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]

            if range_to_remove:
                # Remove validations from the specified range
                dv_to_remove = []
                for dv in sheet.data_validations.dataValidation:
                    if range_to_remove in dv.ranges:
                        dv_to_remove.append(dv)
                for dv in dv_to_remove:
                    sheet.data_validations.dataValidation.remove(dv)
                logger.info(f"Data validation removed from range '{range_to_remove}' in sheet '{sheet_name}'.")
            else:
                # Remove all validations from the sheet
                sheet.data_validations.dataValidation = []
                logger.info(f"All data validations removed from sheet '{sheet_name}'.")

            self.workbook.save(self.file_path)
        except Exception as e:
            logger.error(f"Failed to remove data validation from sheet '{sheet_name}': {e}")
            raise SpreadsheetError(f"Failed to remove data validation from sheet '{sheet_name}': {e}") from e

    # ------------------------------------------------------------------------------
    # 10. Formatting Operations (set_cell_format, apply_conditional_formatting)
    # ------------------------------------------------------------------------------
    
    def set_cell_format(
        self,
        sheet_name: str,
        cell: str,
        style_rules: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Applies style rules to a single cell with default safety.

        Args:
            sheet_name (str): Name of the target sheet.
            cell (str): Cell reference (e.g., 'A1').
            style_rules (Optional[Dict[str, Any]]): Styling rules.
                Example:
                {
                    "font": {"bold": True, "color": "FF0000", "size": 14},
                    "fill": {"fgColor": "FFFF00"}
                }

        Raises:
            SpreadsheetError: If applying the format fails.
        """
        try:
            self._ensure_workbook_loaded()

            sheet = self.workbook[sheet_name]
            target_cell = sheet[cell]

            if style_rules is None:
                # Provide safe defaults if no style rules are provided
                style_rules = {
                    "font": {"bold": False, "color": "000000", "size": 11},
                    "fill": {"fgColor": "FFFFFF"},
                }

            # Apply font styles
            if "font" in style_rules:
                font_defaults = {"bold": False, "color": "000000", "size": 11}
                user_font = style_rules["font"]
                merged_font = {**font_defaults, **user_font}
                target_cell.font = Font(**merged_font)

            # Apply fill styles
            if "fill" in style_rules:
                fill_defaults = {"fgColor": "FFFFFF"}  # White background
                user_fill = style_rules["fill"]
                merged_fill = {**fill_defaults, **user_fill}
                # Ensure patternType is set
                if "patternType" not in merged_fill:
                    merged_fill["patternType"] = "solid"
                target_cell.fill = PatternFill(**merged_fill)

            logger.info(
                f"set_cell_format -> Applied style rules to {cell} in sheet '{sheet_name}': {style_rules}"
            )
            self.workbook.save(self.file_path)

        except Exception as e:
            logger.error(f"Failed to set cell format for {cell} in sheet '{sheet_name}': {e}")
            raise SpreadsheetError(
                f"Failed to set cell format for {cell} in sheet '{sheet_name}': {e}"
            ) from e


    def apply_conditional_formatting(
        self, 
        sheet_name: str, 
        formatting_rules: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Applies conditional formatting to a sheet.

        Args:
            sheet_name (str): Name of the target sheet.
            formatting_rules (Optional[Dict[str, Any]]): Conditional formatting rules.
                Example:
                {
                    "range": "A1:A10",
                    "type": "containsText",
                    "text": "Hello",
                    "font": {"color": "FF0000", "bold": True},
                    "fill": {"fgColor": "FFFF00"}
                }

        Note:
            This is a minimal stub. For full functionality, implement using openpyxl's conditional formatting classes.

        Raises:
            SpreadsheetError: If applying the formatting fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet = self.workbook[sheet_name]

            if formatting_rules is None:
                # Provide a safe "no-op" default
                formatting_rules = {
                    "range": "A1:A1",
                    "type": "containsText",
                    "text": "Default",
                    "font": {"color": "000000"},
                    "fill": {}
                }

            target_range = formatting_rules.get("range", "A1:A1")
            rule_type = formatting_rules.get("type", "containsText")

            # Log the application of conditional formatting
            logger.info(
                f"apply_conditional_formatting -> Stub applying '{rule_type}' rule to '{target_range}' with rules: {formatting_rules}"
            )

            # Save to ensure any changes are persisted
            self.workbook.save(self.file_path)

        except Exception as e:
            logger.error(f"Failed to apply conditional formatting to sheet '{sheet_name}': {e}")
            raise SpreadsheetError(
                f"Failed to apply conditional formatting to sheet '{sheet_name}': {e}"
            ) from e


    # ------------------------------------------------------------------------------
    # 11. Chart and Graphics Operations (create_chart, update_chart, remove_chart)
    # ------------------------------------------------------------------------------
    
    def create_chart(
        self,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        categories_range: Optional[str] = None,
        destination_cell: str = "A10",
        title: Optional[str] = None,
        x_title: Optional[str] = None,
        y_title: Optional[str] = None,
        legend_position: Optional[str] = "r",
        style: Optional[int] = 2,
        show_data_labels: bool = False,
        overlap: Optional[int] = None,
        grouping: Optional[str] = None,
        series_names: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Create and insert a chart into a sheet.

        Args:
            sheet_name (str): Target sheet.
            chart_type (str): 'bar','line','pie','scatter','area','bubble'.
            data_range (str): Numeric data range (e.g. 'B2:D10').
            categories_range (Optional[str]): Labels range (e.g. 'A2:A10').
            destination_cell (str): Top-left cell for chart.
            title (Optional[str]): Chart title.
            x_title (Optional[str]): X-axis title.
            y_title (Optional[str]): Y-axis title.
            legend_position (Optional[str]): 'r','b', etc.
            style (Optional[int]): Chart style index.
            show_data_labels (bool): Show data labels if supported.
            overlap (Optional[int]): Overlap for bar/column.
            grouping (Optional[str]): 'clustered' or 'stacked'.
            series_names (Optional[List[str]]): Titles per series.

        Returns:
            Dict[str,Any]: {'status', 'message', 'result': {'sheet_name', 'chart_type', 'destination_cell'}}

        Raises:
            SpreadsheetError: On invalid parameters or failure.
        """
        try:
            self._ensure_workbook_loaded()

            # 1. Validate the sheet
            if sheet_name not in self.workbook.sheetnames:
                error_msg = f"Sheet '{sheet_name}' does not exist."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            sheet = self.workbook[sheet_name]

            # 2. Parse data_range for numeric data
            try:
                min_col, min_row, max_col, max_row = range_boundaries(data_range)
            except ValueError:
                error_msg = f"Invalid data_range format: '{data_range}'. Expected something like 'B2:D10'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            # 3. Create a Reference for the data
            data_ref = Reference(
                sheet,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row
            )

            # 4. Parse categories_range if provided
            cats_ref = None
            if categories_range:
                try:
                    c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(categories_range)
                    cats_ref = Reference(
                        sheet,
                        min_col=c_min_col,
                        min_row=c_min_row,
                        max_col=c_max_col,
                        max_row=c_max_row
                    )
                except ValueError:
                    error_msg = f"Invalid categories_range format: '{categories_range}'."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)

            # 5. Instantiate the appropriate chart object
            chart_type_lower = chart_type.lower()
            if chart_type_lower == "bar":
                chart = BarChart()
            elif chart_type_lower == "line":
                chart = LineChart()
            elif chart_type_lower == "pie":
                chart = PieChart()
            elif chart_type_lower == "scatter":
                chart = ScatterChart()
            elif chart_type_lower == "area":
                chart = AreaChart()
            elif chart_type_lower == "bubble":
                chart = BubbleChart()
            else:
                error_msg = f"Unsupported chart type '{chart_type}'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            # 6. Apply chart style, titles, etc.
            chart.style = style
            if title:
                chart.title = title
            if x_title:
                chart.x_axis.title = x_title
            if y_title:
                chart.y_axis.title = y_title

            # Optional grouping / overlap for bar/column charts
            if hasattr(chart, "grouping") and grouping:
                chart.grouping = grouping
            if hasattr(chart, "overlap") and overlap is not None:
                chart.overlap = overlap

            # 7. Build series from data_ref
            col_count = (max_col - min_col) + 1
            row_count = (max_row - min_row) + 1

            # If multiple columns => multiple series
            if col_count > 1:
                for idx in range(col_count):
                    series_ref = Reference(
                        sheet,
                        min_col=min_col + idx,
                        max_col=min_col + idx,
                        min_row=min_row,
                        max_row=max_row
                    )
                    series = Series(series_ref, title_from_data=True)
                    if cats_ref:
                        series.categories = cats_ref
                    if series_names and idx < len(series_names):
                        series.title = series_names[idx]
                    chart.series.append(series)
            else:
                # single column => single series
                single_series = Series(data_ref, title_from_data=True)
                if cats_ref:
                    single_series.categories = cats_ref
                if series_names and len(series_names) > 0:
                    single_series.title = series_names[0]
                chart.series.append(single_series)

            # 8. If show_data_labels is True, show data labels (if supported)
            if show_data_labels and hasattr(chart, "dataLabels"):
                chart.dataLabels = DataLabelList(showVal=True)

            # 9. Legend position
            chart.legend.position = legend_position

            # 10. Insert chart into sheet
            try:
                dest_col_letters = ''.join(filter(str.isalpha, destination_cell))
                dest_row_str = ''.join(filter(str.isdigit, destination_cell))

                if not dest_col_letters or not dest_row_str:
                    error_msg = f"Invalid destination cell '{destination_cell}'. Expected format like 'A10'."
                    logger.error(error_msg)
                    raise SpreadsheetError(error_msg)

                # We won't do extensive numeric checks, but let's parse them
                _ = column_index_from_string(dest_col_letters)
                _ = int(dest_row_str)
            except ValueError:
                error_msg = f"Invalid destination cell '{destination_cell}'. Expected format like 'A10'."
                logger.error(error_msg)
                raise SpreadsheetError(error_msg)

            sheet.add_chart(chart, destination_cell)

            # 11. Save the workbook
            self.workbook.save(self.file_path)
            logger.info(f"Created {chart_type} chart at '{destination_cell}' on sheet '{sheet_name}' successfully.")

            return {
                "status": True,
                "message": f"{chart_type.title()} chart created at '{destination_cell}' in sheet '{sheet_name}'.",
                "result": {
                    "sheet_name": sheet_name,
                    "chart_type": chart_type,
                    "destination_cell": destination_cell
                }
            }

        except SpreadsheetError:
            raise
        except Exception as e:
            error_msg = f"Failed to create chart: {e}"
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)


    def update_chart(
        self,
        sheet_name: str,
        chart_title: str,
        new_data_range: Optional[str] = None,
        new_categories_range: Optional[str] = None,
        new_title: Optional[str] = None,
        new_x_title: Optional[str] = None,
        new_y_title: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Update an existing chart’s data or titles.

        Args:
            sheet_name (str): Sheet containing the chart.
            chart_title (str): Current chart.title.
            new_data_range (Optional[str]): New data range.
            new_categories_range (Optional[str]): New category range.
            new_title (Optional[str]): New chart title.
            new_x_title (Optional[str]): New X-axis title.
            new_y_title (Optional[str]): New Y-axis title.

        Returns:
            Dict[str,Any]: {'status','message','result':{'sheet_name','chart_title'}}

        Raises:
            SpreadsheetError: On invalid sheet/chart or failure.
        """
        try:
            self._ensure_workbook_loaded()
            if sheet_name not in self.workbook.sheetnames:
                raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")

            sheet = self.workbook[sheet_name]

            # 1. Find the chart object by title
            target_chart = None
            for obj in sheet._charts:
                # Typically chart.title is a string or None
                if obj.title == chart_title:
                    target_chart = obj
                    break

            if not target_chart:
                raise SpreadsheetError(f"No chart found with title '{chart_title}' in sheet '{sheet_name}'.")

            # 2. If new titles are provided, update them
            if new_title is not None:
                target_chart.title = new_title
            if new_x_title is not None:
                target_chart.x_axis.title = new_x_title
            if new_y_title is not None:
                target_chart.y_axis.title = new_y_title

            # 3. If new_data_range is provided, update references
            if new_data_range:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(new_data_range)
                    data_ref = Reference(
                        sheet,
                        min_col=min_col,
                        min_row=min_row,
                        max_col=max_col,
                        max_row=max_row
                    )
                except ValueError:
                    raise SpreadsheetError(f"Invalid new_data_range format: '{new_data_range}'.")

                # Clear old series; we re-add them
                target_chart.series = []

                # For simplicity, we won't differentiate if multi-column or single-column:
                # If multi-column, each becomes a new series
                col_count = (max_col - min_col) + 1
                for idx in range(col_count):
                    series_ref = Reference(
                        sheet,
                        min_col=min_col + idx,
                        max_col=min_col + idx,
                        min_row=min_row,
                        max_row=max_row
                    )
                    series = Series(series_ref, title_from_data=True)
                    target_chart.series.append(series)

            # 4. If new_categories_range is provided, update the category references
            if new_categories_range:
                try:
                    c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(new_categories_range)
                    cats_ref = Reference(
                        sheet,
                        min_col=c_min_col,
                        min_row=c_min_row,
                        max_col=c_max_col,
                        max_row=c_max_row
                    )
                except ValueError:
                    raise SpreadsheetError(f"Invalid new_categories_range format: '{new_categories_range}'.")

                # Apply categories to every series
                for s in target_chart.series:
                    s.categories = cats_ref

            # 5. Save changes
            self.workbook.save(self.file_path)
            logger.info(f"Chart '{chart_title}' updated successfully on sheet '{sheet_name}'.")

            return {
                "status": True,
                "message": f"Chart '{chart_title}' updated successfully on sheet '{sheet_name}'.",
                "result": {
                    "sheet_name": sheet_name,
                    "chart_title": new_title if new_title else chart_title
                }
            }

        except SpreadsheetError:
            raise
        except Exception as e:
            error_msg = f"Failed to update chart '{chart_title}' on sheet '{sheet_name}': {e}"
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)


    def remove_chart(
        self,
        sheet_name: str,
        chart_title: str
    ) -> Dict[str, Any]:
        """
        Remove a chart by its title.

        Args:
            sheet_name (str): Sheet containing the chart.
            chart_title (str): Title of chart to remove.

        Returns:
            Dict[str,Any]: {'status','message','result':{'sheet_name'}}

        Raises:
            SpreadsheetError: On invalid sheet/chart or failure.
        """
        try:
            self._ensure_workbook_loaded()
            if sheet_name not in self.workbook.sheetnames:
                raise SpreadsheetError(f"Sheet '{sheet_name}' does not exist.")

            sheet = self.workbook[sheet_name]

            target_chart = None
            for obj in sheet._charts:
                if obj.title == chart_title:
                    target_chart = obj
                    break

            if not target_chart:
                raise SpreadsheetError(f"No chart found with title '{chart_title}' in sheet '{sheet_name}'.")

            # Remove this chart from the sheet
            sheet._charts.remove(target_chart)

            self.workbook.save(self.file_path)
            logger.info(f"Chart '{chart_title}' removed successfully from sheet '{sheet_name}'.")

            return {
                "status": True,
                "message": f"Chart '{chart_title}' removed successfully from sheet '{sheet_name}'.",
                "result": {
                    "sheet_name": sheet_name
                }
            }

        except SpreadsheetError:
            raise
        except Exception as e:
            error_msg = f"Failed to remove chart '{chart_title}' on sheet '{sheet_name}': {e}"
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)


    # ------------------------------------------------------------------------------
    # 12. Helper Methods
    # ------------------------------------------------------------------------------

    def _column_index(self, column_letter: str) -> int:
        """
        Converts a column letter (e.g., 'A') to a 0-based index.

        Args:
            column_letter (str): Letter of the column.

        Returns:
            int: 0-based index of the column.

        Raises:
            SpreadsheetError: If the column letter is invalid.
        """
        try:
            index = column_index_from_string(column_letter) - 1  # Convert to 0-based index
            if index < 0:
                raise SpreadsheetError(f"Invalid column letter '{column_letter}'. Must be A or higher.")
            logger.debug(f"Converted column letter '{column_letter}' to index {index}.")
            return index
        except Exception as e:
            logger.error(f"Error converting column letter '{column_letter}' to index: {e}")
            raise SpreadsheetError(f"Error converting column letter '{column_letter}' to index: {e}") from e


    def _parse_range(self, cell_range: str) -> Tuple[Worksheet, str]:
        """
        Parses a cell range string like 'Sheet1!A1:C3' and returns (Worksheet, 'A1:C3').

        Args:
            cell_range (str): Cell range string.

        Returns:
            Tuple[Worksheet, str]: Worksheet object and cell range string.

        Raises:
            SpreadsheetError: If parsing fails.
        """
        try:
            self._ensure_workbook_loaded()
            sheet_part, cells_part = cell_range.split('!')
            sheet = self.workbook[sheet_part]

            # Validate individual cell coordinates
            if ':' in cells_part:
                start_cell, end_cell = cells_part.split(':')
                coordinate_from_string(start_cell)  # Validates start_cell
                coordinate_from_string(end_cell)    # Validates end_cell
            else:
                coordinate_from_string(cells_part)  # Validates single cell

            return sheet, cells_part
        except KeyError:
            error_msg = f"Sheet '{sheet_part}' does not exist."
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)
        except ValueError:
            error_msg = f"Invalid cell range format: '{cell_range}'. Expected format 'SheetName!A1:C3'."
            logger.error(error_msg)
            raise SpreadsheetError(error_msg)
        except Exception as e:
            logger.error(f"Failed to parse range '{cell_range}': {e}")
            raise SpreadsheetError(f"Failed to parse range '{cell_range}': {e}") from e


    def _resolve_column_identifier(
        self,
        sheet_name: str,
        identifier: Union[str, int],
        has_headers: bool = True
    ) -> int:
        """
        Resolves 'identifier' to a 0-based column index.
        
        Args:
            sheet_name (str): Name of the target sheet.
            identifier (Union[str, int]): Could be:
            - A column letter (e.g. "A" or "D").
            - A 1-based column index (e.g. 1, 4).
            - A header name (e.g. "Price", "Product").
            has_headers (bool): Whether to treat row #1 as headers.

        Returns:
            int: 0-based column index suitable for openpyxl operations.
        
        Raises:
            SpreadsheetError: If the sheet or the column cannot be resolved.
        """
        # 1) If integer, interpret as a 1-based column index:
        if isinstance(identifier, int):
            # Convert 1-based to 0-based
            return identifier - 1
        
        # 2) If it “looks like” a letter-based reference, try the built-in method:
        #    e.g. "A", "B", "AA" ...
        try:
            # If this fails, it’ll raise ValueError
            col_idx = column_index_from_string(identifier)  
            return col_idx - 1
        except ValueError:
            pass  # Means 'identifier' was not a valid letter like "A" or "BD"
        
        # 3) Otherwise, assume it’s a "header name"
        #    We'll search row 1 (or whichever is your “header row”)
        #    to find the matching header text.
        sheet = self.workbook[sheet_name]
        
        if not has_headers:
            # If user says "no headers", we can’t interpret "Price" properly
            raise SpreadsheetError(
                f"Requested header '{identifier}' but 'has_headers' is false."
            )
        
        # read row #1 in that sheet
        headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), []))
        if identifier not in headers:
            raise SpreadsheetError(
                f"Header '{identifier}' not found in sheet '{sheet_name}'. "
                f"Available headers: {headers}"
            )
        # find which position it’s at
        col_idx = headers.index(identifier)  # zero-based
        return col_idx


    def _build_condition_func(self, condition_type: str, condition_value: str):
        """
        Builds a condition function based on the condition type and value
        (e.g., 'equals', 'greater_than', etc.).

        Args:
            condition_type (str): Type of condition.
            condition_value (str): Value to compare against.

        Returns:
            Callable: A function that takes a cell value and returns a boolean.

        Raises:
            SpreadsheetError: If the condition type is unsupported or invalid.
        """
        try:
            if condition_type == "equals":
                return lambda x: x == condition_value
            elif condition_type == "greater_than":
                return lambda x: float(x) > float(condition_value) if x is not None else False
            elif condition_type == "less_than":
                return lambda x: float(x) < float(condition_value) if x is not None else False
            elif condition_type == "contains":
                return lambda x: condition_value in str(x) if x is not None else False
            elif condition_type == "startswith":
                return lambda x: str(x).startswith(condition_value) if x is not None else False
            elif condition_type == "endswith":
                return lambda x: str(x).endswith(condition_value) if x is not None else False
            else:
                raise SpreadsheetError(f"Unsupported condition type '{condition_type}'.")
        except ValueError:
            raise SpreadsheetError(
                f"Invalid condition value '{condition_value}' for '{condition_type}'. Must be a number if using > or <."
            )
        except Exception as e:
            logger.error(f"Error building condition function: {e}")
            raise SpreadsheetError(f"Error building condition function: {e}") from e
